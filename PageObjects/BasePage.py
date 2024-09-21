from openpyxl.reader.excel import load_workbook
from selenium.webdriver.support import expected_conditions as EC
import allure
from appium import webdriver
from appium.options.common.base import AppiumOptions
from appium.webdriver.common.appiumby import AppiumBy
from selenium.common import TimeoutException, NoSuchElementException
from selenium.webdriver.support.wait import WebDriverWait
from TestCases.BaseTest import BaseTest

# For W3C actions
from typing import TYPE_CHECKING, List, Optional, Tuple, cast
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.actions import interaction
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.common.actions.mouse_button import MouseButton
from selenium.webdriver.common.actions.pointer_input import PointerInput


if TYPE_CHECKING:
    from appium.webdriver.webdriver import WebDriver


class BasePage(BaseTest):

    def __init__(self,driver,source=None, duration=250):
        self.driver = driver


    def Type(self,text,locator_name,locator_value):
        with allure.step(f"Type text {text} into the {self.ElementsName(locator_name)} Input box"):
            if text:
                element = self.get_element(locator_name, locator_value)
                try:
                    WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((AppiumBy.XPATH, locator_value)))
                    element.click()
                    element.clear()
                    element.send_keys(text)
                    if self.driver.is_keyboard_shown():
                        self.driver.hide_keyboard()
                except Exception as e:
                    assert False, f"{e}"

    def Click(self,locator_name,locator_value):
        with allure.step( f"Click on {self.ElementsName(locator_name)} button or link" ):
            if self.driver.is_keyboard_shown():
                self.driver.hide_keyboard()
            try:
                element = self.get_element(locator_name, locator_value)
                WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((AppiumBy.XPATH, locator_value)))
                element.click()
            except Exception as e:
                assert False, f"{e}"

    def check_display_status_of_element(self,locator_name,locator_value):
        with allure.step(f"Verify {self.ElementsName(locator_name)} webelement is displayed"):
            try:
                WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((AppiumBy.XPATH, locator_value)))
                element = self.get_element(locator_name,locator_value)
                assert element.is_displayed(), f"Element is present on the page."
            except Exception:
                assert False, f"{self.ElementsName(locator_name)} Element is not present on the page."

    def retrieve_element_text(self,locator_name,locator_value):
        element = self.get_element(locator_name,locator_value)
        return element.text


    def selectValueFromComboBox(self, ComboBoxValue, locator_name, locator_value):
        with allure.step( f"Select the {ComboBoxValue} value from {self.ElementsName( locator_name )} comboBox" ):
            try:
                # Click on the combobox to open the dropdown
                self.Click(locator_name,locator_value)

                # Locate and click the desired option in the dropdown
                option_xpath = f"//*[starts-with(text(),'{ComboBoxValue}')]"
                WebDriverWait( self.driver, 10 ).until( EC.presence_of_element_located( (AppiumBy.XPATH, option_xpath) ) )
                self.driver.execute_script( "arguments[0].click();",self.driver.find_element(AppiumBy.XPATH, option_xpath ) )
            except Exception as e:
                return (f"An error occurred: {str(e)}")


    def ReadExcelColumn(SheetName, ColumnName):
        ExcelPath = "../ExcelFiles path"
        # ExcelPath = "C://Users//ahanumantharayappa//PycharmProjects//Aldar_Salesforce_Phase2//ExcelFiles//DamscusWebData.xlsx"

        try:
            # Load the Excel file
            workbook = load_workbook(ExcelPath, data_only=True)
            sheet = workbook[SheetName]

            # Find the column index based on the column name
            col_index = None
            for cell in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                if ColumnName in cell:
                    col_index = cell.index(ColumnName)
                    break

            if col_index is not None:
                # Read values from the specified column
                values = sheet.cell(row=2, column=col_index + 1).value
                return values
            else:
                return f"Column '{ColumnName}' not found in the Excel file."
        except FileNotFoundError:
            return f"File '{ExcelPath}' not found."
        except Exception as e:
            return f"An error occurred: {str(e)}"

        # -----------------------------------------------------

    def Scroll_ByElement(self, locator_name, locator_value):
        for i in range(15):
            x = 950
            try:
                element = self.get_element(locator_name, locator_value)
                value = element.is_displayed()
                if value is True:
                    break
            except NoSuchElementException:
                # swipe(start_x, start_y, end_x, end_y, duration)
                self.driver.swipe(470, 1400, 470, x, 330)
                self.driver.implicitly_wait(1)
                continue

    def Scroll_ByDistance(self,distance=400):
        action = ActionChains(self.driver)
        window_size = self.driver.get_window_size()
        x, y = window_size['width'] / 6, window_size['height'] / 2
        action.w3c_actions = ActionBuilder(self.driver, mouse=PointerInput(interaction.POINTER_TOUCH, 'touch'))
        action.w3c_actions.pointer_action.move_to_location(x, y)
        action.w3c_actions.pointer_action.click_and_hold()
        action.w3c_actions.pointer_action.move_to_location(x, y - distance)
        action.w3c_actions.pointer_action.release()
        action.w3c_actions.perform()

    # --------------------------------------------------------------------------------------------------------------------------
    def get_element(self,locator_name,locator_value):
        element = None
        if locator_name.endswith("_XPATH"):
            # WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((AppiumBy.XPATH, locator_value)))
            element = self.driver.find_element(AppiumBy.XPATH, locator_value)
        elif locator_name.endswith("_NAME"):
            element = self.driver.find_element(AppiumBy.NAME, locator_value)
        elif locator_name.endswith("_CLASS_NAME"):
            element = self.driver.find_element(AppiumBy.CLASS_NAME, locator_value)
        elif locator_name.endswith("_LINK_TEXT"):
            element = self.driver.find_element(AppiumBy.LINK_TEXT, locator_value)
        elif locator_name.endswith("_ID"):
            element = self.driver.find_element(AppiumBy.ID, locator_value)
        elif locator_name.endswith("_CSS_SELECTOR"):
            element = self.driver.find_element(AppiumBy.CSS_SELECTOR, locator_value)
        return element

    def ElementsName(self,locator_name):
        b = locator_name.split( "_" )
        return b[1]